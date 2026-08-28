#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证投注簿星级列（原始单元格）"""
from openpyxl import load_workbook

wb = load_workbook("/data/data/com.termux/files/home/storage/shared/Documents/投注簿.xlsx")
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if not any(row):
        continue
    cells = [str(c)[:30] if c is not None else '' for c in row]
    line = ' | '.join(cells).rstrip(' |')
    if line.strip():
        print(line)
