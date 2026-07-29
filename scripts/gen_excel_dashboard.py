#!/usr/bin/env python3
"""
生成 Excel 清新看板 (两页)
  Page 1: 比赛看板 — 今日/未来比赛核心数据
  Page 2: 历史相同亚盘 — 各盘口历史统计
"""

import json, os, sqlite3, sys
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter

# ========== 配置 ==========
PROJ_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_PATH = os.path.join(PROJ_DIR, "docs", "data", "results.json")
DB_PATH = os.path.join(PROJ_DIR, "data", "football.db")
OUTPUT_DIR = os.path.join(PROJ_DIR, "docs", "data")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "dashboard.xlsx")
TODAY = datetime.now().strftime("%Y-%m-%d")

# ========== 样式 ==========
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if r % 2 == 0:
                cell.fill = LIGHT_GRAY


def auto_width(ws, max_col, max_row, min_width=8, max_width=28):
    for col in range(1, max_col + 1):
        longest = min_width
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                # Estimate: CJK chars ~2 width, others ~1
                s = str(val)
                w = sum(2 if ord(c) > 127 else 1 for c in s)
                longest = max(longest, min(w + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = longest


# ========== 加载数据 ==========
def load_data():
    with open(RESULTS_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    matches = raw.get("matches", [])
    return matches


def safe_float(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def ev_str(best_value):
    """Format EV value"""
    if not best_value:
        return ""
    ev = best_value.get("ev", 0)
    if ev is None:
        return ""
    try:
        ev = float(ev)
    except:
        return ""
    if ev > 0:
        return f"+{ev:.1%}"
    return f"{ev:.1%}"


def kelly_str(best_value):
    if not best_value:
        return ""
    k = best_value.get("kelly", 0)
    if k is None:
        return ""
    try:
        k = float(k)
    except:
        return ""
    if k <= 0:
        return ""
    return f"{k:.1%}"


# ========== Page 1: 比赛看板 ==========
def build_sheet1(ws, matches):
    """未来比赛核心看板"""
    ws.title = "比赛看板"

    # Title
    ws.cell(row=1, column=1, value=f"⚽ 足球看板 — {TODAY}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)

    # Headers
    headers = [
        "序号", "联赛", "开赛时间", "主队",
        "让球", "客队",
        "主胜", "平", "客胜",
        "模型方向", "概率",
        "EV值", "凯利"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    # Filter future matches (no score or score is placeholder)
    future = [
        m for m in matches
        if not m.get("score") or m["score"] in ["? - ?", "", "vs", None, "0-0"]
    ]

    # Sort by match_time
    future.sort(key=lambda m: m.get("match_time", m.get("date", "")))

    row = 4
    for idx, m in enumerate(future, 1):
        # Skip obviously already-played matches (score is present and not placeholder)
        score = m.get("score", "")
        if score and score not in ["? - ?", "", "vs", None, "0-0"]:
            # Check if score is a real result (contains digits on both sides)
            parts = str(score).split("-")
            if len(parts) == 2 and parts[0].strip().isdigit() and parts[1].strip().isdigit():
                if int(parts[0]) > 0 or int(parts[1]) > 0:
                    continue  # real result, skip

        best_val = m.get("best_value")
        pred_cn = m.get("prediction_cn", m.get("model_prediction_cn", ""))
        pred_prob = m.get("prediction_prob", m.get("model_prediction_prob", 0))
        ah_text = m.get("ah_handicap_text", "")
        if not ah_text and m.get("ah_handicap") is not None:
            h = float(m.get("ah_handicap", 0))
            ah_text = f"{h:+.1f}" if h != 0 else "平手"

        # Skip matches more than 7 days in the future
        mt = m.get("match_time", m.get("date", ""))
        try:
            mt_dt = datetime.strptime(mt[:10], "%Y-%m-%d") if mt else None
            if mt_dt and (mt_dt - datetime.now()).days > 7:
                continue
        except:
            pass

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=m.get("event", m.get("league", "")))
        # Time
        raw_time = m.get("match_time", m.get("date", ""))
        if len(raw_time) >= 16:
            ws.cell(row=row, column=3, value=raw_time[5:16])
        else:
            ws.cell(row=row, column=3, value=raw_time)
        ws.cell(row=row, column=4, value=m.get("home_team", ""))
        ws.cell(row=row, column=5, value=ah_text)

        # Color handicap text based on value
        ah_v = m.get("ah_handicap")
        if ah_v is not None:
            try:
                if float(ah_v) < 0:
                    ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="E2EFDA")  # greenish - home favored
                elif float(ah_v) > 0:
                    ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FCE4D6")  # orangeish - away favored
            except:
                pass

        ws.cell(row=row, column=6, value=m.get("away_team", ""))

        # Odds
        ws.cell(row=row, column=7, value=safe_float(m.get("odds_win"), None))
        ws.cell(row=row, column=8, value=safe_float(m.get("odds_draw"), None))
        ws.cell(row=row, column=9, value=safe_float(m.get("odds_loss"), None))
        for c in [7, 8, 9]:
            if ws.cell(row=row, column=c).value and ws.cell(row=row, column=c).value > 0:
                ws.cell(row=row, column=c).number_format = "0.00"

        # Model prediction
        ws.cell(row=row, column=10, value=pred_cn)
        prob_val = safe_float(pred_prob, None)
        if prob_val:
            ws.cell(row=row, column=11, value=prob_val)
            ws.cell(row=row, column=11).number_format = "0.0%"
            # Color: green if high confidence
            if prob_val >= 0.45:
                ws.cell(row=row, column=10).fill = GREEN_FILL
                ws.cell(row=row, column=11).fill = GREEN_FILL
            elif prob_val >= 0.35:
                ws.cell(row=row, column=10).fill = YELLOW_FILL
                ws.cell(row=row, column=11).fill = YELLOW_FILL

        # EV
        ev = best_val.get("ev", 0) if best_val else 0
        if ev is not None:
            ev = safe_float(ev, 0)
            ws.cell(row=row, column=12, value=ev)
            ws.cell(row=row, column=12).number_format = "+0.0%;-0.0%"
            if ev > 0.1:
                ws.cell(row=row, column=12).fill = GREEN_FILL
            elif ev < -0.05:
                ws.cell(row=row, column=12).fill = RED_FILL

        # Kelly
        kelly_val = best_val.get("kelly", 0) if best_val else 0
        if kelly_val:
            kelly_val = safe_float(kelly_val, 0)
            if kelly_val > 0:
                ws.cell(row=row, column=13, value=kelly_val)
                ws.cell(row=row, column=13).number_format = "0.0%"

        row += 1

    end_row = row - 1
    style_body(ws, 4, end_row, len(headers))
    auto_width(ws, len(headers), end_row)

    # Summary row
    row += 1
    ws.cell(row=row, column=1, value=f"共 {end_row - 3} 场比赛").font = Font(
        name="微软雅黑", size=10, italic=True, color="666666"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)


# ========== Page 2: 历史相同亚盘 ==========
def build_sheet2(ws, matches):
    """各盘口历史统计"""
    ws.title = "历史相同亚盘"

    # Title
    ws.cell(row=1, column=1, value=f"📊 历史相同亚盘统计 — {TODAY}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # Collect completed matches with handicap + score
    completed_ah = []
    for m in matches:
        score = m.get("score", "")
        ah = m.get("ah_handicap")
        if not score or score in ["? - ?", "", "vs", None]:
            continue
        if ah is None:
            continue
        # Parse score
        parts = str(score).split("-")
        if len(parts) != 2:
            continue
        try:
            hg = int(parts[0].strip())
            ag = int(parts[1].strip())
        except:
            continue
        ah_text = m.get("ah_handicap_text", f"{float(ah):+.1f}")
        completed_ah.append({
            "ah": float(ah),
            "ah_text": ah_text,
            "home": m.get("home_team", ""),
            "away": m.get("away_team", ""),
            "hg": hg,
            "ag": ag,
            "result": "home" if hg > ag else ("draw" if hg == ag else "away"),
        })

    if not completed_ah:
        ws.cell(row=3, column=1, value="暂无已完赛且有亚盘数据的比赛").font = BODY_FONT
        return

    # Also gather from DB (poisson_predictions with reference_score)
    db_data = []
    try:
        db = sqlite3.connect(DB_PATH)
        pp_rows = db.execute('''
            SELECT reference_score, actual_outcome, league, home_team, away_team
            FROM poisson_predictions
            WHERE reference_score IS NOT NULL AND reference_score != ''
              AND actual_outcome IS NOT NULL AND actual_outcome != ''
            ORDER BY date DESC
            LIMIT 500
        ''').fetchall()
        for r in pp_rows:
            score_str = r[0]
            parts = str(score_str).split("-")
            if len(parts) == 2:
                try:
                    hg = int(parts[0].strip())
                    ag = int(parts[1].strip())
                    db_data.append({
                        "home": r[3],
                        "away": r[4],
                        "league": r[2],
                        "hg": hg,
                        "ag": ag,
                        "result": "home" if hg > ag else ("draw" if hg == ag else "away"),
                    })
                except:
                    pass
        db.close()
    except Exception as e:
        print(f"DB query error: {e}", file=sys.stderr)

    # Section A: By Asian handicap line
    ws.cell(row=3, column=1, value="按盘口统计").font = Font(
        name="微软雅黑", size=12, bold=True, color="1F4E79"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    headers_a = [
        "盘口", "场次", "主胜", "平局", "客胜",
        "主胜率", "平率", "客胜率"
    ]
    for i, h in enumerate(headers_a, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers_a))

    # Group by handicap text
    ah_groups = defaultdict(list)
    for m in completed_ah:
        ah_groups[m["ah_text"]].append(m)

    # Sort by frequency
    sorted_ahs = sorted(ah_groups.items(), key=lambda x: len(x[1]), reverse=True)

    row = 5
    for ah_text, group in sorted_ahs:
        total = len(group)
        home_w = sum(1 for g in group if g["result"] == "home")
        draws = sum(1 for g in group if g["result"] == "draw")
        away_w = sum(1 for g in group if g["result"] == "away")

        ws.cell(row=row, column=1, value=ah_text)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=home_w)
        ws.cell(row=row, column=4, value=draws)
        ws.cell(row=row, column=5, value=away_w)
        ws.cell(row=row, column=6, value=home_w / total if total else 0)
        ws.cell(row=row, column=6).number_format = "0.0%"
        ws.cell(row=row, column=7, value=draws / total if total else 0)
        ws.cell(row=row, column=7).number_format = "0.0%"
        ws.cell(row=row, column=8, value=away_w / total if total else 0)
        ws.cell(row=row, column=8).number_format = "0.0%"

        # Color the highest rate
        rates = [(ws.cell(row=row, column=6), home_w / total),
                 (ws.cell(row=row, column=7), draws / total),
                 (ws.cell(row=row, column=8), away_w / total)]
        max_cell, max_rate = max(rates, key=lambda x: x[1])
        if max_rate > 0.4 and total >= 3:
            max_cell.fill = GREEN_FILL

        row += 1

    style_body(ws, 5, row - 1, len(headers_a))
    end_a = row - 1

    # Section B: Overall league + handicap cross analysis (if enough data)
    row += 2
    ws.cell(row=row, column=1, value="全场统计").font = Font(
        name="微软雅黑", size=12, bold=True, color="1F4E79"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    total_all = len(completed_ah)
    hw_all = sum(1 for m in completed_ah if m["result"] == "home")
    d_all = sum(1 for m in completed_ah if m["result"] == "draw")
    aw_all = sum(1 for m in completed_ah if m["result"] == "away")

    summary_data = [
        ("已完赛有亚盘场次", total_all),
        ("主胜", f"{hw_all} ({hw_all/total_all:.1%})" if total_all else 0),
        ("平局", f"{d_all} ({d_all/total_all:.1%})" if total_all else 0),
        ("客胜", f"{aw_all} ({aw_all/total_all:.1%})" if total_all else 0),
        ("", ""),
        ("模型推荐胜率统计（可用于校准）", ""),
    ]

    # Count prediction hit rate
    hits = sum(1 for m in matches
               if m.get("prediction_cn") and m.get("hit") and "✓" in str(m.get("hit", "")))
    total_preds = sum(1 for m in matches
                      if m.get("prediction_cn") and m.get("hit"))
    if total_preds:
        summary_data.append(("模型推荐准确率", f"{hits}/{total_preds} ({hits/total_preds:.1%})"))

    for label, val in summary_data:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=val).font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    auto_width(ws, 8, max(end_a, row))


# ========== 主函数 ==========
def main():
    matches = load_data()
    if not matches:
        print("❌ results.json 为空或无法加载")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    # Page 1
    ws1 = wb.active
    build_sheet1(ws1, matches)

    # Page 2
    ws2 = wb.create_sheet()
    build_sheet2(ws2, matches)

    wb.save(OUTPUT_FILE)
    print(f"✅ 看板已生成: {OUTPUT_FILE}")

    # Print brief summary
    future = [m for m in matches
              if not m.get("score") or m["score"] in ["? - ?", "", "vs", None]]
    print(f"   比赛看板: {sum(1 for _ in filter(None, [True]))} 场 (包含今日~7日内的比赛)")
    print(f"   历史亚盘: {sum(1 for m in matches if m.get('ah_handicap') is not None and m.get('score') and m['score'] not in ['? - ?', '', 'vs', None])} 场已完成")

    return OUTPUT_FILE


if __name__ == "__main__":
    main()
