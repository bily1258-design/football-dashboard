#!/usr/bin/env python3
"""生成「双模型一致+TS反向」列表 → xlsx，存到手机共享 Documents 供 WPS 打开存云文档。

用法: python3 gen_dual_consensus_xlsx.py [YYYY-MM-DD] [HH:MM]
输出: ~/storage/shared/Documents/双模型一致TS反向_YYYYMMDD.xlsx
"""
import json, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_PATH = os.path.join(PROJECT_DIR, 'docs', 'data', 'results.json')
OUT_DIR = os.path.expanduser('~/storage/shared/Documents')

def direction(w, d, l):
    m = max(w, d, l)
    return '主' if m == w else ('平' if m == d else '客')

def main():
    today = datetime.now().strftime('%Y-%m-%d')
    cutoff_time = '18:00'
    if len(sys.argv) > 1:
        today = sys.argv[1]
    if len(sys.argv) > 2:
        cutoff_time = sys.argv[2]
    cutoff = f'{today} {cutoff_time}'

    with open(RESULTS_PATH) as f:
        data = json.load(f)
    matches = data.get('matches', [])

    hits = []
    for m in matches:
        mt = m.get('match_time', '')
        if mt < cutoff:
            continue
        md = direction(m.get('model_win', 0), m.get('model_draw', 0), m.get('model_loss', 0))
        ld = direction(m.get('lgbm_win', 0), m.get('lgbm_draw', 0), m.get('lgbm_loss', 0))
        td = direction(m.get('ts_win', 0), m.get('ts_draw', 0), m.get('ts_loss', 0))
        comb = md + ld + td
        if comb not in ('客客主', '主主客'):
            continue
        hits.append((m, md, ld, td, comb))
    hits.sort(key=lambda x: x[0]['match_time'])

    wb = Workbook()
    ws = wb.active
    ws.title = '双模型一致TS反向'

    headers = ['方向', '比赛时间', '主队', '客队', '跟单方向', '比分', '模型(主/平/客)', 'LGBM(主/平/客)', 'TS(主/平/客)', '欧赔(主/平/客)', '亚盘']
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='C00000')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    dir_fill = {'客客主': PatternFill('solid', fgColor='FCE4EC'),  # 红底（跟客胜）
                '主主客': PatternFill('solid', fgColor='E3F2FD')}  # 蓝底（跟主胜）

    for m, md, ld, td, comb in hits:
        follow = '客胜' if comb == '客客主' else '主胜'
        pct = lambda w, d, l: f'{w:.0%}/{d:.0%}/{l:.0%}'
        row = [comb, m.get('match_time', '')[11:16], m['home_team'], m['away_team'], follow,
               m.get('score') or '-',
               pct(m.get('model_win', 0), m.get('model_draw', 0), m.get('model_loss', 0)),
               pct(m.get('lgbm_win', 0), m.get('lgbm_draw', 0), m.get('lgbm_loss', 0)),
               pct(m.get('ts_win', 0), m.get('ts_draw', 0), m.get('ts_loss', 0)),
               f"{m.get('odds_win','-')}/{m.get('odds_draw','-')}/{m.get('odds_loss','-')}",
               m.get('ah_handicap_text') or '-']
        ws.append(row)
        r = ws.max_row
        fill = dir_fill[comb]
        for c in ws[r]:
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.fill = fill

    widths = [8, 10, 22, 22, 10, 10, 20, 20, 20, 22, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    os.makedirs(OUT_DIR, exist_ok=True)
    date_str = today.replace('-', '')
    out = os.path.join(OUT_DIR, f'双模型一致TS反向_{date_str}.xlsx')
    wb.save(out)

    # 控制台/消息输出（与 list_dual_consensus_ts_reverse.py 同款详细格式）
    print(f'📋 「双模型一致+TS反向」筛选 [{cutoff} 之后] 共 {len(hits)} 场')
    print('=' * 90)
    for m, md, ld, td, comb in hits:
        h, a = m['home_team'], m['away_team']
        mt = m['match_time']
        sc = m.get('score') or '-'
        ow, od, ol = m.get('odds_win'), m.get('odds_draw'), m.get('odds_loss')
        odds = f'{ow}/{od}/{ol}' if ow else '-'
        ah = m.get('ah_handicap_text') or '-'
        mw, mdr, ml = m.get('model_win'), m.get('model_draw'), m.get('model_loss')
        lw, ldr, ll = m.get('lgbm_win'), m.get('lgbm_draw'), m.get('lgbm_loss')
        tw, tdr, tl = m.get('ts_win'), m.get('ts_draw'), m.get('ts_loss')
        print(f'[{comb}] {mt}  {h} vs {a}  比分:{sc}')
        print(f'      模型 {mw:.0%}/{mdr:.0%}/{ml:.0%} | LGBM {lw:.0%}/{ldr:.0%}/{ll:.0%} | TS {tw:.0%}/{tdr:.0%}/{tl:.0%}')
        print(f'      欧赔 {odds} | 亚盘 {ah}')
        print('-' * 90)
    print(f'✅ 已生成 {out}（{len(hits)} 场），手机 Documents 目录，WPS 打开后可存云文档')

if __name__ == '__main__':
    main()
