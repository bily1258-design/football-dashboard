#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成「投注簿.xlsx」= Excel 版 today_picks.md（今日推荐清单）。

用法: python3 gen_ledger_xlsx.py
输入: docs/today_picks.md (away_value_picks.py --md 生成, 12:30 cron 已更新)
输出: ~/storage/shared/Documents/投注簿.xlsx (固定文件名, 每日覆盖更新)
内容:
  Sheet1 今日推荐: 按档位(🎯甜点区/②高置信/③客客客)逐场: 时间/联赛/对阵/方向/概率/EV/赔率/避雷
  Sheet2 避雷汇总: 🚫 全部避雷场次, 慎跟
"""
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MD_PATH = os.path.join(PROJECT_DIR, 'docs', 'today_picks.md')
OUT_DIR = os.path.expanduser('~/storage/shared/Documents')

# 高命中率联赛 (results.json 1877场回测, n>=20 且方向命中率>=60%, 2026-08-21)
# 芬甲70% 国际友谊赛67.3% 日职联65% 智利甲64.3% 欧冠杯63.9% 挪甲63.3%
# 挪超62.5% 丹麦超61.9% 欧罗巴杯61.9% 英联杯61.1% — 联赛列绿字加粗
# 2026-08-22: 同步 away_value_picks.py 简体联赛名(清单已转简体, 繁体集合匹配不上)
HIGH_HIT_LEAGUES = {'芬甲', '国际友谊赛', '日职联', '智利甲', '欧冠杯',
                    '挪甲', '挪超', '丹麦超', '欧罗巴杯', '英联杯'}

# 样式
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='2F5597')
SWEET_FILL = PatternFill('solid', fgColor='C6EFCE')   # 甜点区 绿
CONF_FILL = PatternFill('solid', fgColor='DDEBF7')    # 高置信 蓝
KKK_FILL = PatternFill('solid', fgColor='FFF2CC')     # 客客客 黄
AVOID_FILL = PatternFill('solid', fgColor='FFC7CE')   # 避雷 红
HIGH_HIT_FONT = Font(color='008000', bold=True)        # 高命中率联赛 绿字加粗
RED_FONT = Font(color='FF0000', bold=True)              # 红线 红字加粗
STAR_FONT = Font(color='E67E22', bold=True)             # 星级 橙字加粗
TITLE_FONT = Font(bold=True, size=14, color='2F5597')
SECTION_FONT = Font(bold=True, size=12, color='404040')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEC_RE = re.compile(r'^(①|②|③|⚠️)')
MATCH_RE = re.compile(
    r'^(\d{2}-\d{2})\s+(\d{2}:\d{2})\s+\[([^\]]+)\]\s+(.+?)\s*(→(主|平|客))?\s*(🎯|★)?\s*((?:🚫|⚠️⚡).*)?$')
AVOID_ITEM_RE = re.compile(
    r'^\s*(\d{2}-\d{2})\s+(\d{2}:\d{2})\s+\[([^\]]+)\]\s+(.+?)\s+🚫(.+)$')
PROB_RE = re.compile(
    r'^(主|客|平)概率:\s*model\s*(\d+)%\s*\|\s*LGBM\s*(\d+)%\s*(?:\|\s*EV\s*([\d.]+))?\s*(?:\|\s*TS\s*(主|客|平)\s*(\d+)%)?')
HKJC_RE = re.compile(
    r'^HKJC(客|主|平)胜\s*([\d.]+)\s*\|\s*'
    r'(?:(?:模型概率\s*(\d+)%\s*\|\s*)?LGBM(客|主|平)概率\s*(\d+)%|\s*模型概率\s*(\d+)%)'
    r'\s*(?:\|\s*EV\s*([\d.]+))?\s*(?:\|\s*TS(平|主|客)\s*(\d+)%)?')
ODDS_RE = re.compile(
    r'^平博\s+初/即:\s*([\d./\-]+)\s*→\s*([\d./\-]+)\s*\|\s*HKJC\s+初/即:\s*([\d./\-]+)\s*→\s*([\d./\-]+)')

# ===== 2026-08-28 星级评估 (★★★★) =====
# ① 平博升水 + HKJC(掉水或不变) → 直接★★
# ② 有★赛事(md行尾★, 赔率<2.0且TS平<25%) → 加1★
# ③ HKJC升水 → 不碰 (红线 7.9%, 命中率红线)
# ④ 平博掉水 → 不作红线 (误杀, 17.1%与基准持平)
# ⑤ 高命中联赛🟢 → 再加1★
DIR_IDX = {'主': 0, '平': 1, '客': 2}


def parse_triple(s):
    """'1.93/3.62/3.51' → [1.93, 3.62, 3.51]; '-'/非法 → None"""
    try:
        parts = [float(x) for x in s.split('/')]
        if len(parts) == 3:
            return parts
    except (ValueError, AttributeError):
        pass
    return None


def calc_stars(d, p0, p1, h0, h1, has_star, league):
    """返回 (星级字符串, 红线标记)
    d=推荐方向(主/平/客); p0/p1=平博初/即三元组; h0/h1=HKJC初/即三元组
    """
    idx = DIR_IDX.get(d)
    if idx is None:
        return '', ''
    stars = 0
    red = ''
    pt0, pt1 = parse_triple(p0), parse_triple(p1)
    ht0, ht1 = parse_triple(h0), parse_triple(h1)
    pin_up = pt0 and pt1 and pt1[idx] > pt0[idx]          # 平博升水
    hk_up = ht0 and ht1 and ht1[idx] > ht0[idx]           # HKJC升水
    hk_down = ht0 and ht1 and ht1[idx] < ht0[idx]         # HKJC掉水
    hk_same = ht0 and ht1 and abs(ht1[idx] - ht0[idx]) < 1e-9
    if pin_up and (hk_down or hk_same):                   # ①
        stars += 2
    if has_star:                                          # ②
        stars += 1
    if league in HIGH_HIT_LEAGUES:                        # ⑤ 🟢
        stars += 1
    if hk_up:                                             # ③ 红线
        red = '🚫HKJC升水·不碰'
    return '★' * stars, red


def parse_md(path):
    """解析 today_picks.md → (sections, avoids)
    sections: [(档位, 标题, [场次]), ...]  场次=dict
    """
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()
    sections = []
    avoids = []
    cur_sec = None
    cur_match = None

    def flush_match():
        nonlocal cur_match
        if cur_match is not None and cur_sec is not None:
            cur_sec[2].append(cur_match)
        cur_match = None

    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        m_sec = SEC_RE.match(s)
        if m_sec:                               # 档位标题
            flush_match()
            cur_sec = [m_sec.group(1), s, []]
            sections.append(cur_sec)
            continue
        if cur_sec and cur_sec[0] == '⚠️':      # ⚠️档内: 避雷明细行
            m = AVOID_ITEM_RE.match(ln)
            if m:
                avoids.append({'date': m.group(1), 'time': m.group(2),
                               'league': m.group(3), 'teams': m.group(4).strip(),
                               'reason': m.group(5).strip()})
                continue
        m = MATCH_RE.match(s)                   # 场次行
        if m:
            flush_match()
            cur_match = {
                'date': m.group(1), 'time': m.group(2), 'league': m.group(3),
                'teams': m.group(4).strip(), 'dir': m.group(6) or '',
                'star': m.group(7) or '', 'avoid': m.group(8) or '',
                'prob_line': '', 'hkjc_line': '', 'odds_line': '',
            }
            continue
        if cur_match is not None:               # 属性行
            if s.startswith('平博'):
                cur_match['odds_line'] = s
            elif s.startswith('HKJC'):
                cur_match['hkjc_line'] = s
            elif s.startswith(('主概率', '客概率', '平概率')):
                cur_match['prob_line'] = s
    flush_match()
    return sections, avoids


def build_rows(sections):
    """sections → 表格行列表 (含档位标题行标记)"""
    rows = []
    for idx, title, matches in sections:
        if idx == '⚠️':
            continue
        default_dir = '客' if idx in ('①', '②') else ('主' if idx == '③' else '')
        for mt in matches:
            d = mt['dir'] or default_dir
            star = mt['star'] == '★'
            hk_odds, mdl, lgbm, ev = '', '', '', ''
            tsd = tsp = ''
            if mt['hkjc_line']:
                m = HKJC_RE.search(mt['hkjc_line'])
                if m:
                    hk_odds = m.group(2)
                    mdl = m.group(3) or m.group(6) or ''   # 模型概率(共存或单独)
                    lgbm = m.group(5) or ''                # LGBM概率
                    ev = m.group(7) or ''
                    if m.group(8):                         # TS平/主/客 在HKJC行(②③档)
                        tsd, tsp = m.group(8), m.group(9)
            if mt['prob_line']:
                m = PROB_RE.search(mt['prob_line'])
                if m:
                    d, mdl, lgbm, ev, tsd, tsp = (m.group(1), m.group(2),
                                                  m.group(3), m.group(4) or '',
                                                  m.group(5) or '', m.group(6) or '')
            p0 = p1 = h0 = h1 = ''
            if mt['odds_line']:
                m = ODDS_RE.search(mt['odds_line'])
                if m:
                    p0, p1, h0, h1 = m.group(1), m.group(2), m.group(3), m.group(4)
                    if not hk_odds:                       # 高置信档无HKJC行, 取即盘对应推荐方向赔率(主→主胜/客→客胜/平→平赔)
                        try:
                            parts = h1.split('/')
                            if len(parts) == 3:
                                hk_odds = parts[0] if d == '主' else parts[2] if d == '客' else parts[1]
                        except (IndexError, AttributeError):
                            pass
            league = mt['league'].replace('🟢', '')
            stars_str, red = calc_stars(d, p0, p1, h0, h1, star, league)
            rows.append({
                'sec': idx, 'date': mt['date'], 'time': mt['time'],
                'league': league, 'teams': mt['teams'],
                'dir': f"→{d}",
                'stars': stars_str, 'red': red,
                'mdl': int(mdl) if mdl else '',
                'lgbm': int(lgbm) if lgbm else '',
                'ev': float(ev) if ev else '',
                'ts': f"{tsd}{tsp}%" if tsd else '',
                'hk_odds': float(hk_odds) if hk_odds else '',
                'p_odds': f"{p0} → {p1}" if p0 else '',
                'h_odds': f"{h0} → {h1}" if h0 else '',
                'avoid': mt['avoid'].replace('🚫避雷', '🚫').replace('⚠️⚡避雷', '⚠️⚡') if mt['avoid'] else '',
            })
    return rows


def main():
    if not os.path.exists(MD_PATH):
        print(f'❌ 未找到 {MD_PATH}, 先跑 fetch_and_push.sh / away_value_picks.py --md')
        return 1
    sections, avoids = parse_md(MD_PATH)
    rows = build_rows(sections)
    wb = Workbook()

    # ─── Sheet1 今日推荐 ─────────────────────────
    ws = wb.active
    ws.title = '今日推荐'
    ws['A1'] = '📋 今日投注参考（Excel 版 today_picks.md）'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = ('生成时间: %s  |  完整清单: https://bily1258-design.github.io/football-dashboard/today_picks.md'
                % datetime.now().strftime('%Y-%m-%d %H:%M'))
    ws['A2'].font = Font(size=10, color='808080')

    headers = ['档位', '日期', '时间', '联赛', '对阵', '方向', '星级', 'Model%', 'LGBM%', 'EV',
               'TS', 'HKJC赔率', '平博 初→即', 'HKJC 初→即', '避雷']
    widths = [7, 8, 8, 13, 30, 8, 9, 7, 7, 7, 11, 9, 26, 26, 22]
    SEC_FILL = {'①': CONF_FILL, '②': KKK_FILL, '③': KKK_FILL}
    n_secs = 0
    for idx, title, matches in sections:
        if idx == '⚠️':
            continue
        n_secs += 1

    row = 4
    for idx, title, matches in sections:
        if idx == '⚠️':
            continue
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        row += 1
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        row += 1
        fill = SEC_FILL.get(idx)
        for r in [r for r in rows if r['sec'] == idx]:
            vals = [idx, r['date'], r['time'], r['league'], r['teams'], r['dir'],
                    r['stars'], r['mdl'], r['lgbm'], r['ev'], r['ts'], r['hk_odds'],
                    r['p_odds'], r['h_odds'], ' '.join(x for x in (r['avoid'], r['red']) if x)]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = BORDER
                if fill and not r['avoid']:
                    cell.fill = fill
                if ci == 4 and r['league'] in HIGH_HIT_LEAGUES:   # 高命中率联赛 绿字加粗
                    cell.font = HIGH_HIT_FONT
                if ci == 7 and r['red']:                           # 红线 红色加粗
                    cell.font = RED_FONT
                if ci == 7 and r['stars'] and not r['red']:        # 星级 橙色
                    cell.font = STAR_FONT
                if ci in (1, 2, 3, 6, 7, 8, 9, 10, 11):
                    cell.alignment = Alignment(horizontal='center')
            row += 1
        row += 1

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = 'A5'

    # ─── Sheet2 避雷汇总 ─────────────────────────
    ws2 = wb.create_sheet('避雷汇总')
    ws2['A1'] = '⚠️🚫 避雷场次（历史败率 87-93%，慎跟）'
    ws2['A1'].font = SECTION_FONT
    hdrs2 = ['日期', '时间', '联赛', '对阵', '避雷原因']
    for ci, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    rr = 3
    for av in avoids:
        for ci, v in enumerate([av['date'], av['time'], av['league'], av['teams'], av['reason']], 1):
            cell = ws2.cell(row=rr, column=ci, value=v)
            cell.border = BORDER
            cell.fill = AVOID_FILL
        rr += 1
    for ci, w in enumerate([8, 8, 14, 32, 30], 1):
        ws2.column_dimensions[chr(64 + ci)].width = w
    ws2.freeze_panes = 'A3'

    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, '投注簿.xlsx')
    wb.save(out_path)
    n_matches = len(rows)
    print(f'✅ 已生成: {out_path}')
    print(f'   推荐 {n_matches} 场 ({n_secs} 档) + 避雷 {len(avoids)} 场')


if __name__ == '__main__':
    main()
