#!/usr/bin/env python3
"""
从500.com抓取北单26079期数据，生成Excel清新看板（单表）
合并让球胜平负 + 胜负过关两页数据
"""

import re, os, sys
from datetime import datetime
from urllib.request import Request, urlopen
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJ_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJ_DIR, "docs", "data")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "beidan_26079_dashboard.xlsx")
TODAY = datetime.now().strftime("%Y-%m-%d")

# ===== Styles =====
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
HOME_FAV_FILL = PatternFill("solid", fgColor="E2EFDA")
AWAY_FAV_FILL = PatternFill("solid", fgColor="FCE4D6")


def fetch_page(playid, expect=26079):
    url = f"https://zx.500.com/zqdc/saiguo.php?playid={playid}&expect={expect}"
    req = Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Linux; Android 13; K) AppleWebKit/537.36",
        "Referer": "https://zx.500.com/zqdc/",
    })
    with urlopen(req, timeout=15) as resp:
        raw = resp.read()
    try:
        return raw.decode("gb2312")
    except:
        return raw.decode("gbk", errors="replace")


def parse_matches(text, playid):
    """Parse table rows from 500.com page"""
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL)
    matches = {}
    for r in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', r, re.DOTALL)
        if len(cells) < 6:
            continue
        vals = []
        for c in cells:
            clean = re.sub(r'<[^>]+>', ' ', c)
            clean = re.sub(r'&nbsp;|\s+', ' ', clean).strip()
            vals.append(clean)
        first = vals[0].strip()
        if not first.isdigit():
            continue
        num = int(first)

        if playid == 3:
            # 让球胜平负: cells [0]num [1]league [2]time [3]home [4]handicap [5]away [6]score [9]胜% [11]平% [13]负%
            if len(vals) >= 14:
                probs = [v.rstrip("%") for v in [vals[9], vals[11], vals[13]] if "%" in v]
                matches[num] = {
                    "num": num, "league": vals[1], "time": vals[2],
                    "home": vals[3], "handicap": vals[4], "away": vals[5],
                    "score": vals[6], "result": "",
                    "h_prob3": int(probs[0]) if len(probs) > 0 else -1,
                    "d_prob3": int(probs[1]) if len(probs) > 1 else -1,
                    "a_prob3": int(probs[2]) if len(probs) > 2 else -1,
                }
        else:
            # 胜负过关: [0]num [1]league [2]time [3]home [4]handicap [5]away [6]score [8]亚盘 [10]胜% [12]平% [14]负%
            if len(vals) >= 15:
                probs = [v.rstrip("%") for v in [vals[10], vals[12], vals[14]] if "%" in v]
                matches[num] = {
                    "num": num, "league": vals[1], "time": vals[2],
                    "home": vals[3], "handicap": vals[4], "away": vals[5],
                    "score": vals[6], "result": "",
                    "ah_desc": vals[8],
                    "h_prob0": int(probs[0]) if len(probs) > 0 else -1,
                    "d_prob0": int(probs[1]) if len(probs) > 1 else -1,
                    "a_prob0": int(probs[2]) if len(probs) > 2 else -1,
                }
    return matches


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if r % 2 == 0:
                cell.fill = LIGHT_GRAY


def auto_width(ws, max_col, max_row, min_w=8, max_w=30):
    for col in range(1, max_col + 1):
        longest = min_w
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                s = str(val)
                w = sum(2 if ord(c) > 127 else 1 for c in s)
                longest = max(longest, min(w + 2, max_w))
        ws.column_dimensions[get_column_letter(col)].width = longest


def prob_cell(ws, row, col, prob_val):
    """Write probability with color coding"""
    cell = ws.cell(row=row, column=col)
    if prob_val >= 0:
        cell.value = f"{prob_val}%"
        if prob_val >= 45:
            cell.fill = GREEN_FILL
        elif prob_val >= 35:
            cell.fill = YELLOW_FILL


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("🔄 抓取北单26079期数据...")

    m3 = parse_matches(fetch_page(3, 26079), 3)
    m0 = parse_matches(fetch_page(0, 26079), 0)
    print(f"   ✓ 让球胜平负: {len(m3)} 场  胜负过关: {len(m0)} 场")

    # Merge both data sources by match number
    all_nums = sorted(set(list(m3.keys()) + list(m0.keys())))
    print(f"   ✓ 合并后共 {len(all_nums)} 场")

    wb = Workbook()
    ws = wb.active
    ws.title = "北单26079期"

    ws.cell(row=1, column=1, value=f"⚽ 北京单场26079期 比赛看板 — {TODAY}").font = TITLE_FONT

    headers = [
        "序号", "联赛", "开赛时间", "主队", "让球(胜平负)", "客队",
        "亚盘(胜负过关)", "历史概率% (主胜)", "历史概率% (平)", "历史概率% (客胜)"
    ]
    max_col = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, max_col)

    row = 4
    for num in all_nums:
        d3 = m3.get(num, {})
        d0 = m0.get(num, {})

        league = d3.get("league", d0.get("league", ""))
        time_val = d3.get("time", d0.get("time", ""))
        home = d3.get("home", d0.get("home", ""))
        away = d3.get("away", d0.get("away", ""))
        hdcp = d3.get("handicap", d0.get("handicap", ""))
        ah_desc = d0.get("ah_desc", "")

        # Use playid=0 probs if available, else playid=3
        if "h_prob0" in d0 and d0["h_prob0"] >= 0:
            h_prob, d_prob, a_prob = d0["h_prob0"], d0["d_prob0"], d0["a_prob0"]
        else:
            h_prob, d_prob, a_prob = d3.get("h_prob3", -1), d3.get("d_prob3", -1), d3.get("a_prob3", -1)

        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=league)
        ws.cell(row=row, column=3, value=time_val)
        ws.cell(row=row, column=4, value=home)

        # Handicap with color
        hdcp_cell = ws.cell(row=row, column=5, value=hdcp)
        try:
            hnum = int(hdcp)
            if hnum < 0:
                hdcp_cell.fill = HOME_FAV_FILL
            elif hnum > 0:
                hdcp_cell.fill = AWAY_FAV_FILL
        except:
            pass

        ws.cell(row=row, column=6, value=away)
        ws.cell(row=row, column=7, value=ah_desc)

        prob_cell(ws, row, 8, h_prob)
        prob_cell(ws, row, 9, d_prob)
        prob_cell(ws, row, 10, a_prob)

        row += 1

    end_row = row - 1
    style_body(ws, 4, end_row, max_col)
    auto_width(ws, max_col, end_row)

    # Summary
    row += 1
    ws.cell(row=row, column=1, value=f"共 {end_row - 3} 场比赛").font = Font(
        name="微软雅黑", size=10, italic=True, color="666666"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # Legend
    row += 2
    ws.cell(row=row, column=1, value="图例：").font = BOLD_FONT
    row += 1
    for fill, desc in [
        (HOME_FAV_FILL, "让球为负（主队让球）"),
        (AWAY_FAV_FILL, "让球为正（客队让球）"),
        (GREEN_FILL, "历史概率≥45%"),
        (YELLOW_FILL, "历史概率≥35%"),
    ]:
        c = ws.cell(row=row, column=1, value="  ")
        c.fill = fill
        c.border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).font = BODY_FONT
        row += 1

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 看板已生成: {OUTPUT_FILE}")
    print(f"\n📋 预览 (前5场):")
    for num in all_nums[:5]:
        d3, d0 = m3.get(num, {}), m0.get(num, {})
        l = d3.get("league", d0.get("league", ""))
        t = d3.get("time", d0.get("time", ""))
        h = d3.get("home", d0.get("home", ""))
        a = d3.get("away", d0.get("away", ""))
        hc = d3.get("handicap", d0.get("handicap", ""))
        ah = d0.get("ah_desc", "")
        hp = d0.get("h_prob0", d3.get("h_prob3", -1))
        dp = d0.get("d_prob0", d3.get("d_prob3", -1))
        ap = d0.get("a_prob0", d3.get("a_prob3", -1))
        print(f"   {num}. {l} {t} {h}({hc}){a} [{ah}] 胜{hp}%平{dp}%负{ap}%")

    return OUTPUT_FILE


if __name__ == "__main__":
    main()
