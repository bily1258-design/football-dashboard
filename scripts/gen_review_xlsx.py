#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成「复盘.xlsx」= Excel 版 推荐清单·赛果回填复盘.md（昨日清单逐场回填比分）。

用法: python3 gen_review_xlsx.py
输入: docs/推荐清单·赛果回填复盘.md (fetch_and_push.sh 内 gen_daily_review.py 生成, 固定名每日覆盖)
输出: ~/storage/shared/Documents/复盘.xlsx (固定文件名, 每日覆盖更新)
内容:
  Sheet1 复盘明细: 按档位(①甜点区/②高置信/③客客客)逐场: 方向/比分/结果/概率/EV/赔率/避雷/盈亏
  Sheet2 档位汇总: 每档 场数/完赛/命中/命中率/净盈亏
"""
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MD_PATH = os.path.join(PROJECT_DIR, 'docs', '推荐清单·赛果回填复盘.md')
OUT_DIR = os.path.expanduser('~/storage/shared/Documents')

# 高命中率联赛 (简体, 与 gen_ledger_xlsx.py 同步, 绿字加粗)
# 2026-08-31: 复盘md联赛名是简体+🟢(如丹麦超🟢), 原繁体集合匹配不上致+1★失效, 同步简体
HIGH_HIT_LEAGUES = {'芬甲', '国际友谊赛', '日职联', '智利甲', '欧冠杯',
                    '挪甲', '挪超', '丹麦超', '欧罗巴杯', '英联杯',
                    '爱甲', '奥甲', 'NCAL Cup', '捷甲', '非女杯', '西甲',
                    'SPL', '荷甲', '英乙', '澳维超', '瑞典超'}

# 样式
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='2F5597')
SWEET_FILL = PatternFill('solid', fgColor='C6EFCE')   # 甜点区 绿
CONF_FILL = PatternFill('solid', fgColor='DDEBF7')    # 高置信 蓝
KKK_FILL = PatternFill('solid', fgColor='FFF2CC')     # 客客客 黄
HIT_FONT = Font(color='006100', bold=True)            # 命中 深绿加粗
MISS_FONT = Font(color='9C0006', bold=True)           # 未中 深红加粗
PEND_FONT = Font(color='808080')                      # 未收录 灰
HIGH_HIT_FONT = Font(color='008000', bold=True)        # 高命中率联赛 绿字加粗
RED_FONT = Font(color='FF0000', bold=True)              # 红线 红字加粗
STAR_FONT = Font(color='E67E22', bold=True)             # 星级 橙字加粗
TITLE_FONT = Font(bold=True, size=14, color='2F5597')
SECTION_FONT = Font(bold=True, size=12, color='404040')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEC_RE = re.compile(r'^(①|②|③|⚠️)')
MATCH_RE = re.compile(
    r'^(\d{2}-\d{2})\s+(\d{2}:\d{2})\s+\[([^\]]+)\]\s+(.+?)\s*(→(主|平|客))?\s*(🎯|★)?\s*((?:🚫|⚠️⚡).*)?$')
AVOID_ITEM_RE = re.compile(
    r'^\s*(\d{2}-\d{2})\s+(\d{2}:\d{2})\s+\[([^\]]+)\]\s+(.+?)\s+🚫(.+)$')
PROB_RE = re.compile(
    r'^(主|客|平)概率:\s*model\s*(\d+)%\s*\|\s*LGBM\s*(\d+)%\s*(?:\|\s*EV\s*([\d.]+))?\s*(?:\|\s*TS\s*(主|客|平)\s*(\d+)%)?')
HKJC_RE = re.compile(
    r'^HKJC(客|主|平)胜\s*([\d.]+)\s*(?:\|\s*模型概率\s*(\d+)%)?\s*(?:\|\s*LGBM(客|主|平)概率\s*(\d+)%)?\s*(?:\|\s*EV\s*([\d.]+))?\s*(?:\|\s*TS(平|主|客)\s*(\d+)%)?')
ODDS_RE = re.compile(
    r'^平博\s+初/即:\s*([\d./\-]+)\s*→\s*([\d./\-]+)\s*\|\s*HKJC\s+初/即:\s*([\d./\-]+)\s*→\s*([\d./\-]+)')
RESULT_RE = re.compile(r'\|\s*实际:\s*(\d+)[-:](\d+)\s*(✓|✘)\s*$')
PEND_RE = re.compile(r'\|\s*⏳\s*未收录\s*$')


def parse_md(path):
    """解析复盘 md → (sections, avoids)
    sections: [(档位, 标题, [场次]), ...]  场次=dict(含 result 字段)
    """
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()
    sections = []
    avoids = []
    cur_sec = None
    cur_match = None

    def flush_match():
        nonlocal cur_match
        if cur_match is not None and cur_sec is not None:
            cur_sec[2].append(cur_match)
        cur_match = None

    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        # 结果后缀 ( | 实际: 2-2 ✓ / | ⏳ 未收录) 从行尾剥离
        result = None
        for rp in (RESULT_RE, PEND_RE):
            m = rp.search(s)
            if m:
                if rp is RESULT_RE:
                    result = {'score': f"{m.group(1)}-{m.group(2)}",
                              'mark': m.group(3)}
                else:
                    result = {'score': None, 'mark': '⏳'}
                s = s[:m.start()].strip()
                break
        m_sec = SEC_RE.match(s)
        if m_sec:                               # 档位标题
            flush_match()
            cur_sec = [m_sec.group(1), s, []]
            sections.append(cur_sec)
            continue
        if cur_sec and cur_sec[0] == '⚠️':      # ⚠️档内: 避雷明细行
            m = AVOID_ITEM_RE.match(ln)
            if m:
                avoids.append({'date': m.group(1), 'time': m.group(2),
                               'league': m.group(3), 'teams': m.group(4).strip(),
                               'reason': m.group(5).strip()})
                continue
        m = MATCH_RE.match(s)                   # 场次行
        if m:
            flush_match()
            cur_match = {
                'date': m.group(1), 'time': m.group(2), 'league': m.group(3),
                'teams': m.group(4).strip(), 'dir': m.group(6) or '',
                'star': m.group(7) or '', 'avoid': m.group(8) or '',
                'result': result,
                'prob_line': '', 'hkjc_line': '', 'odds_line': '',
            }
            continue
        if cur_match is not None:               # 属性行
            if s.startswith('平博'):
                cur_match['odds_line'] = s
            elif s.startswith('HKJC'):
                cur_match['hkjc_line'] = s
            elif s.startswith(('主概率', '客概率', '平概率')):
                cur_match['prob_line'] = s
    flush_match()
    return sections, avoids


# ===== 2026-08-31 星级评估 (对齐投注簿 gen_ledger_xlsx.py calc_stars) =====
DIR_IDX = {'主': 0, '平': 1, '客': 2}


def parse_triple(s):
    """'1.93/3.62/3.51' → [1.93, 3.62, 3.51]; '-'/非法 → None"""
    try:
        parts = [float(x) for x in s.split('/')]
        if len(parts) == 3:
            return parts
    except (ValueError, AttributeError):
        pass
    return None


def calc_stars(d, p0, p1, h0, h1, has_star, league):
    """返回 (星级字符串, 红线标记)。与投注簿口径一致:
    ① 平博升水+HKJC(掉水/不变)→+2★  ② 有★(赔率<2.0且TS平<25%)→+1★
    ⑤ 高命中联赛🟢→+1★  ③ HKJC升水→红线🚫
    """
    idx = DIR_IDX.get(d)
    if idx is None:
        return '', ''
    stars = 0
    red = ''
    pt0, pt1 = parse_triple(p0), parse_triple(p1)
    ht0, ht1 = parse_triple(h0), parse_triple(h1)
    pin_up = pt0 and pt1 and pt1[idx] > pt0[idx]
    hk_up = ht0 and ht1 and ht1[idx] > ht0[idx]
    hk_down = ht0 and ht1 and ht1[idx] < ht0[idx]
    hk_same = ht0 and ht1 and abs(ht1[idx] - ht0[idx]) < 1e-9
    if pin_up and (hk_down or hk_same):
        stars += 2
    if has_star:
        stars += 1
    if league in HIGH_HIT_LEAGUES:
        stars += 1
    if hk_up:
        red = '🚫HKJC升水·不碰'
    return '★' * stars, red


def build_rows(sections, avoid_teams=None):
    """sections → 表格行列表 (含结果/盈亏)
    avoid_teams: 避雷场次 teams 集合 — 2026-08-31 用户拍板: 按场次全局过滤(跨档位),
    同一场在任一档带🚫/⚠️⚡即所有档位都过滤, 避免②③档干净条目混入复盘明细.
    """
    avoid_teams = avoid_teams or set()
    rows = []
    for idx, title, matches in sections:
        if idx == '⚠️':
            continue
        default_dir = '客' if idx in ('①', '②') else ('主' if idx == '③' else '')
        for mt in matches:
            if mt['teams'] in avoid_teams:   # 避雷场次全档位过滤
                continue
            d = mt['dir'] or default_dir
            star = mt['star'] == '★'
            hk_odds, mdl, lgbm, ev = '', '', '', ''
            tsd = tsp = ''
            if mt['hkjc_line']:
                m = HKJC_RE.search(mt['hkjc_line'])
                if m:
                    hk_odds = m.group(2)
                    if m.group(3):                       # 甜点区/高置信: 模型概率
                        mdl = m.group(3)
                        ev = m.group(6) or ''
                        if not m.group(4):               # 无LGBM时模型概率即主值
                            pass
                    if m.group(4):                       # LGBM概率(与模型并存或单独)
                        lgbm = m.group(5)
                        ev = m.group(6) or ev or ''
                    if m.group(8):                       # ②③档TS嵌在HKJC行尾
                        tsd, tsp = m.group(7), m.group(8)
            if mt['prob_line']:
                m = PROB_RE.search(mt['prob_line'])
                if m:
                    d, mdl, lgbm, ev, tsd, tsp = (m.group(1), m.group(2),
                                                  m.group(3), m.group(4) or '',
                                                  m.group(5) or '', m.group(6) or '')
            p0 = p1 = h0 = h1 = ''
            if mt['odds_line']:
                m = ODDS_RE.search(mt['odds_line'])
                if m:
                    p0, p1, h0, h1 = m.group(1), m.group(2), m.group(3), m.group(4)
                    if not hk_odds:                       # 高置信档无HKJC行, 取即盘客胜
                        try:
                            hk_odds = h1.split('/')[2]
                        except (IndexError, AttributeError):
                            pass
            # 盈亏: 与 gen_daily_review.py 口径一致 —
            #   高置信档: HKJC 即盘方向赔率; 甜点区/客客客: HKJC客胜
            pnl = ''
            res = mt.get('result') or {}
            if res.get('mark') in ('✓', '✘'):
                use_odds = hk_odds
                if h1 and d in ('主', '平', '客'):
                    side_idx = {'主': 3, '平': 4, '客': 5}[d]
                    try:
                        v = h1.split('/')[side_idx - 3]
                        if v != '-':
                            use_odds = v
                    except (IndexError, AttributeError):
                        pass
                if use_odds:
                    try:
                        pnl = round(float(use_odds) - 1.0, 2) if res['mark'] == '✓' else -1.0
                    except (ValueError, TypeError):
                        pnl = ''
            league = mt['league'].replace('🟢', '')
            stars_str, red = calc_stars(d, p0, p1, h0, h1, star, league)
            rows.append({
                'sec': idx, 'date': mt['date'], 'time': mt['time'],
                'league': league, 'teams': mt['teams'],
                'dir': f"→{d}",
                'stars': stars_str, 'red': red,
                'score': res.get('score') or '',
                'mark': res.get('mark') or '',
                'mdl': int(mdl) if mdl else '',
                'lgbm': int(lgbm) if lgbm else '',
                'ev': float(ev) if ev else '',
                'ts': f"{tsd}{tsp}%" if tsd else '',
                'hk_odds': float(hk_odds) if hk_odds else '',
                'p_odds': f"{p0} → {p1}" if p0 else '',
                'h_odds': f"{h0} → {h1}" if h0 else '',
                'avoid': mt['avoid'].replace('🚫避雷', '🚫').replace('⚠️⚡避雷', '⚠️⚡') if mt['avoid'] else '',
                'pnl': pnl,
            })
    return rows


def main():
    if not os.path.exists(MD_PATH):
        print(f'❌ 未找到 {MD_PATH}, 先跑 fetch_and_push.sh 生成复盘')
        return 1
    sections, avoids = parse_md(MD_PATH)
    # 2026-08-31 用户拍板: 按场次全局过滤(跨档位)。避雷teams集合 = 避雷汇总 ∪ 档位内带标记场次,
    # 避免避雷汇总漏收(如安特卫普带⚠️⚡但汇总未收录)导致标记场次混入复盘明细.
    avoid_teams = {av['teams'] for av in avoids}
    for _, _, matches in sections:
        for mt in matches:
            if mt.get('avoid'):
                avoid_teams.add(mt['teams'])
    rows = build_rows(sections, avoid_teams)
    wb = Workbook()

    # ─── Sheet1 复盘明细 ─────────────────────────
    ws = wb.active
    ws.title = '复盘明细'
    ws['A1'] = '📋 推荐清单 · 赛果回填复盘（Excel 版）'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = ('生成时间: %s  |  完整清单: https://bily1258-design.github.io/football-dashboard/推荐清单·赛果回填复盘.md'
                % datetime.now().strftime('%Y-%m-%d %H:%M'))
    ws['A2'].font = Font(size=10, color='808080')

    headers = ['档位', '日期', '时间', '联赛', '对阵', '方向', '星级',
               '比分', '结果', 'Model%', 'LGBM%', 'EV', 'TS', 'HKJC赔率',
               '平博 初→即', 'HKJC 初→即', '避雷', '盈亏']
    widths = [7, 8, 8, 13, 30, 8, 9, 8, 7, 7, 7, 7, 11, 9, 26, 26, 22, 8]
    SEC_FILL = {'①': SWEET_FILL, '②': CONF_FILL, '③': KKK_FILL}
    MARK_FONT = {'✓': HIT_FONT, '✘': MISS_FONT, '⏳': PEND_FONT}

    row = 4
    n_total = n_done = n_hit = 0
    sec_stats = {}
    for idx, title, matches in sections:
        if idx == '⚠️':
            continue
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        row += 1
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        row += 1
        fill = SEC_FILL.get(idx)
        st = sec_stats.setdefault(idx, {'n': 0, 'done': 0, 'hit': 0, 'pnl': 0.0})
        for r in [r for r in rows if r['sec'] == idx]:
            vals = [idx, r['date'], r['time'], r['league'], r['teams'], r['dir'],
                    r['stars'], r['score'], r['mark'], r['mdl'], r['lgbm'], r['ev'],
                    r['ts'], r['hk_odds'], r['p_odds'], r['h_odds'],
                    ' '.join(x for x in (r['avoid'], r['red']) if x), r['pnl']]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = BORDER
                if fill:
                    cell.fill = fill
                if ci == 4 and r['league'] in HIGH_HIT_LEAGUES:   # 高命中率联赛 绿字加粗
                    cell.font = HIGH_HIT_FONT
                if ci == 7 and r['red']:                           # 星级列红线 红色加粗
                    cell.font = RED_FONT
                if ci == 7 and r['stars'] and not r['red']:        # 星级 橙色
                    cell.font = STAR_FONT
                if ci == 9 and r['mark']:                          # 结果列 ✓绿 ✘红 ⏳灰
                    cell.font = MARK_FONT.get(r['mark'], PEND_FONT)
                if ci in (1, 2, 3, 6, 7, 8, 9, 11, 12, 18):
                    cell.alignment = Alignment(horizontal='center')
            n_total += 1
            st['n'] += 1
            if r['mark'] in ('✓', '✘'):
                n_done += 1
                st['done'] += 1
                if r['mark'] == '✓':
                    n_hit += 1
                    st['hit'] += 1
            if r['pnl'] != '':
                st['pnl'] += r['pnl']
            row += 1
        row += 1

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = 'A5'

    # ─── Sheet2 档位汇总 ─────────────────────────
    ws2 = wb.create_sheet('档位汇总')
    ws2['A1'] = '📊 各档位命中汇总'
    ws2['A1'].font = SECTION_FONT
    hdrs2 = ['档位', '场数', '完赛', '命中', '命中率', '净盈亏']
    for ci, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    rr = 3
    SEC_NAME = {'①': '甜点区客胜', '②': '高置信方向', '③': '客客客'}
    for idx in ('①', '②', '③'):
        if idx not in sec_stats:
            continue
        st = sec_stats[idx]
        pct = f"{st['hit'] / st['done'] * 100:.0f}%" if st['done'] else '-'
        for ci, v in enumerate([SEC_NAME[idx], st['n'], st['done'], st['hit'],
                                pct, round(st['pnl'], 2)], 1):
            cell = ws2.cell(row=rr, column=ci, value=v)
            cell.border = BORDER
            if ci in (2, 3, 4, 6):
                cell.alignment = Alignment(horizontal='center')
        rr += 1
    # 合计行
    pct_all = f"{n_hit / n_done * 100:.0f}%" if n_done else '-'
    pnl_all = round(sum(st['pnl'] for st in sec_stats.values()), 2)
    for ci, v in enumerate(['合计', n_total, n_done, n_hit, pct_all, pnl_all], 1):
        cell = ws2.cell(row=rr, column=ci, value=v)
        cell.font = Font(bold=True)
        cell.border = BORDER
        if ci in (2, 3, 4, 6):
            cell.alignment = Alignment(horizontal='center')
    for ci, w in enumerate([12, 8, 8, 8, 10, 10], 1):
        ws2.column_dimensions[chr(64 + ci)].width = w

    # ─── Sheet3 避雷汇总 ─────────────────────────
    ws3 = wb.create_sheet('避雷汇总')
    ws3['A1'] = '⚠️🚫 避雷场次（历史败率 87-93%，慎跟）'
    ws3['A1'].font = SECTION_FONT
    hdrs3 = ['日期', '时间', '联赛', '对阵', '避雷原因']
    for ci, h in enumerate(hdrs3, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    rr = 3
    for av in avoids:
        for ci, v in enumerate([av['date'], av['time'], av['league'], av['teams'],
                                av['reason']], 1):
            cell = ws3.cell(row=rr, column=ci, value=v)
            cell.border = BORDER
            cell.fill = PatternFill('solid', fgColor='FFC7CE')
        rr += 1
    for ci, w in enumerate([8, 8, 14, 32, 30], 1):
        ws3.column_dimensions[chr(64 + ci)].width = w
    ws3.freeze_panes = 'A3'

    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, '复盘.xlsx')
    wb.save(out_path)
    pct = f"{n_hit / n_done * 100:.0f}%" if n_done else '-'
    print(f'✅ 已生成: {out_path}')
    print(f'   复盘 {n_total} 场 / 完赛 {n_done} / 命中 {n_hit} ({pct}) / '
          f'净盈亏 {round(sum(st["pnl"] for st in sec_stats.values()), 2):+.2f}')


if __name__ == '__main__':
    main()
