#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证投注簿绿字联赛(高命中)是否生效"""
import openpyxl, sys
sys.path.insert(0, 'scripts')

wb = openpyxl.load_workbook('/data/data/com.termux/files/home/storage/shared/Documents/投注簿.xlsx')
ws = wb.active

# 找到表头行
hdr_row = None
for row in ws.iter_rows(min_row=1, max_row=10, max_col=16):
    for cell in row:
        if cell.value == '联赛':
            hdr_row = cell.row
            break
    if hdr_row: break

print("表头行:", hdr_row)
green = {}
for row in ws.iter_rows(min_row=hdr_row+1, max_col=16):
    league_cell = row[3]   # 第4列 = 联赛
    if league_cell.value is None: continue
    lv = str(league_cell.value)
    font = league_cell.font
    if font.color and font.color.rgb and '008000' in str(font.color.rgb).upper():
        green[lv] = green.get(lv, 0) + 1
print("绿字联赛:", green)
print("绿字总场次:", sum(green.values()))
